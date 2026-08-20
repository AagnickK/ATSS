import os
from flask import Blueprint, render_template, redirect, url_for, request, flash, session, jsonify, send_file
from flask_login import login_user, logout_user, login_required, current_user
from flask_wtf.csrf import generate_csrf
from werkzeug.utils import secure_filename
from extensions import db, bcrypt
from models import User, Faculty, Subject, Division, Room, Allocation, Timetable
from config import DAYS, MORNING_SLOTS, GENERAL_SLOTS, DESIGNATIONS, SHIFTS, SUBJECT_TYPES, ROOM_TYPES, SESSION_TYPES, ODD_SEMESTERS, EVEN_SEMESTERS
import otp as otp_module

# ── Auth Blueprint ────────────────────────────────────────────────────────────
auth = Blueprint('auth', __name__)

@auth.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user = User.query.filter_by(username=request.form['username']).first()
        if user and bcrypt.check_password_hash(user.password, request.form['password']):
            login_user(user)
            return redirect(url_for('main.dashboard'))
        flash('Invalid credentials.', 'error')
    return render_template('login.html')

@auth.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        if not otp_module.is_configured():
            flash('Gmail not configured in .env', 'error')
            return render_template('register.html')
        username = request.form['username']
        email    = request.form['email']
        password = request.form['password']
        if User.query.filter_by(username=username).first():
            flash('Username already exists.', 'error')
            return render_template('register.html')
        if User.query.filter_by(email=email).first():
            flash('Email already registered.', 'error')
            return render_template('register.html')
        code = otp_module.generate_otp()
        try:
            otp_module.send_otp(email, code)
        except Exception as e:
            flash(f'Failed to send OTP: {e}', 'error')
            return render_template('register.html')
        session['pending_user'] = {
            'username': username,
            'email':    email,
            'password': bcrypt.generate_password_hash(password).decode('utf-8'),
        }
        session['otp_code']  = code
        session['otp_email'] = email
        return redirect(url_for('auth.verify_otp'))
    return render_template('register.html')

@auth.route('/verify-otp', methods=['GET', 'POST'])
def verify_otp():
    email = session.get('otp_email', '')
    if not email:
        return redirect(url_for('auth.register'))
    if request.method == 'POST':
        if request.form.get('otp', '').strip() == session.get('otp_code'):
            pending = session.pop('pending_user', None)
            session.pop('otp_code', None)
            if pending:
                db.session.add(User(
                    username=pending['username'],
                    email=pending['email'],
                    password=pending['password'],
                ))
                db.session.commit()
                flash('Account created! Please log in.', 'success')
                return redirect(url_for('auth.login'))
        else:
            flash('Incorrect OTP.', 'error')
    return render_template('verify_otp.html', email=email)

@auth.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('auth.login'))


# ── Main Blueprint ────────────────────────────────────────────────────────────
main = Blueprint('main', __name__)

@main.route('/')
def index():
    return redirect(url_for('auth.login'))

@main.route('/dashboard')
@login_required
def dashboard():
    stats = {
        'faculty':   Faculty.query.count(),
        'subjects':  Subject.query.count(),
        'divisions': Division.query.count(),
        'rooms':     Room.query.count(),
        'entries':   Timetable.query.count(),
    }
    return render_template('dashboard.html', stats=stats)


# ── Faculty Blueprint ─────────────────────────────────────────────────────────
faculty_bp = Blueprint('faculty_bp', __name__)

@faculty_bp.route('/faculty')
@login_required
def faculty_list():
    faculty = Faculty.query.order_by(Faculty.name).all()
    return render_template('faculty.html', faculty=faculty,
                           designations=DESIGNATIONS, shifts=SHIFTS)

@faculty_bp.route('/faculty/add', methods=['POST'])
@login_required
def faculty_add():
    db.session.add(Faculty(
        name        = request.form['name'],
        designation = request.form['designation'],
        department  = request.form['department'],
        shift       = request.form['shift'],
        max_hours   = int(request.form.get('max_hours', 18)),
        email       = request.form.get('email', ''),
    ))
    db.session.commit()
    flash('Faculty added.', 'success')
    return redirect(url_for('faculty_bp.faculty_list'))

@faculty_bp.route('/faculty/delete/<int:fid>', methods=['POST'])
@login_required
def faculty_delete(fid):
    Faculty.query.filter_by(id=fid).delete()
    db.session.commit()
    flash('Faculty removed.', 'success')
    return redirect(url_for('faculty_bp.faculty_list'))

@faculty_bp.route('/faculty/template')
@login_required
def faculty_template():
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Faculty Import'
    headers = [
        'faculty_id', 'faculty_name', 'designation', 'department', 'shift',
        'max_hours', 'email', 'subject_name', 'course', 'semester',
        'type', 'lecture_hours', 'lab_hours', 'division', 'students'
    ]
    hdr_fill = PatternFill('solid', fgColor='FF9000')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='18181B')
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = max(len(h) + 4, 14)
    # Two sample rows
    samples = [
        ['FAC-1001', 'Dr. A. Sharma', 'Associate Professor', 'IT', 'Morning', 14,
         'sharma@college.edu', 'Data Structures', 'B.Tech', 3, 'Theory', 3, 0, 'A', 60],
        ['FAC-1002', 'Prof. B. Mehta', 'Regular Faculty', 'IT', 'General', 18,
         'mehta@college.edu', 'DBMS Lab', 'B.Tech', 4, 'Lab', 0, 2, 'B', 60],
    ]
    for row in samples:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='faculty_import_template.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@faculty_bp.route('/faculty/import', methods=['POST'])
@login_required
def faculty_import():
    from importer import import_faculty_excel
    f = request.files.get('excel')
    if not f:
        flash('No file selected.', 'error')
        return redirect(url_for('faculty_bp.faculty_list'))
    path = os.path.join('data', secure_filename(f.filename))
    f.save(path)
    added = import_faculty_excel(path)
    flash(f'Imported {added} faculty records.', 'success')
    return redirect(url_for('faculty_bp.faculty_list'))


# ── Subjects Blueprint ────────────────────────────────────────────────────────
subject_bp = Blueprint('subject_bp', __name__)

@subject_bp.route('/subjects')
@login_required
def subject_list():
    subjects = Subject.query.order_by(Subject.course, Subject.semester).all()
    return render_template('subjects.html', subjects=subjects, types=SUBJECT_TYPES)

@subject_bp.route('/subjects/add', methods=['POST'])
@login_required
def subject_add():
    db.session.add(Subject(
        subject_name  = request.form['subject_name'],
        course        = request.form['course'],
        semester      = int(request.form['semester']),
        type          = request.form['type'],
        lecture_hours = int(request.form.get('lecture_hours', 0)),
        lab_hours     = int(request.form.get('lab_hours', 0)),
    ))
    db.session.commit()
    flash('Subject added.', 'success')
    return redirect(url_for('subject_bp.subject_list'))

@subject_bp.route('/subjects/delete/<int:sid>', methods=['POST'])
@login_required
def subject_delete(sid):
    Subject.query.filter_by(id=sid).delete()
    db.session.commit()
    flash('Subject removed.', 'success')
    return redirect(url_for('subject_bp.subject_list'))


# ── Rooms Blueprint ───────────────────────────────────────────────────────────
room_bp = Blueprint('room_bp', __name__)

@room_bp.route('/rooms')
@login_required
def room_list():
    rooms = Room.query.order_by(Room.room_no).all()
    return render_template('classroom.html', rooms=rooms, types=ROOM_TYPES)

@room_bp.route('/rooms/add', methods=['POST'])
@login_required
def room_add():
    db.session.add(Room(
        room_no  = request.form['room_no'],
        type     = request.form['type'],
        capacity = int(request.form.get('capacity', 30)),
        building = request.form.get('building', ''),
    ))
    db.session.commit()
    flash('Room added.', 'success')
    return redirect(url_for('room_bp.room_list'))

@room_bp.route('/rooms/delete/<int:rid>', methods=['POST'])
@login_required
def room_delete(rid):
    Room.query.filter_by(id=rid).delete()
    db.session.commit()
    flash('Room removed.', 'success')
    return redirect(url_for('room_bp.room_list'))


# ── Timetable Blueprint ───────────────────────────────────────────────────────
tt_bp = Blueprint('tt_bp', __name__)

@tt_bp.route('/timetable')
@login_required
def timetable_view():
    entries    = Timetable.query.all()
    divisions  = Division.query.order_by(Division.course, Division.semester, Division.division).all()
    faculty    = Faculty.query.order_by(Faculty.name).all()
    sel_div    = request.args.get('division_id', type=int)
    sel_fac    = request.args.get('faculty_id',  type=int)
    sel_session = request.args.get('session_type', '')   # 'Odd' | 'Even' | ''

    if sel_session:
        sems = ODD_SEMESTERS if sel_session == 'Odd' else EVEN_SEMESTERS
        entries = [e for e in entries if e.division and e.division.semester in sems]
    if sel_div:
        entries = [e for e in entries if e.division_id == sel_div]
    if sel_fac:
        entries = [e for e in entries if e.faculty_id == sel_fac]

    return render_template('timetable.html',
                           entries=entries, days=DAYS,
                           morning_slots=MORNING_SLOTS, general_slots=GENERAL_SLOTS,
                           divisions=divisions, faculty=faculty,
                           sel_div=sel_div, sel_fac=sel_fac,
                           sel_session=sel_session, session_types=SESSION_TYPES)

@tt_bp.route('/timetable/generate', methods=['POST'])
@login_required
def timetable_generate():
    from scheduler import generate_timetable
    session_type = request.form.get('session_type', '')   # 'Odd' | 'Even' | ''

    # Clear unlocked entries for the selected session (or all if none selected)
    if session_type:
        sems = ODD_SEMESTERS if session_type == 'Odd' else EVEN_SEMESTERS
        locked_ids = {e.id for e in Timetable.query.filter_by(locked=True).all()}
        to_delete = [
            e for e in Timetable.query.filter_by(locked=False).all()
            if e.division and e.division.semester in sems
        ]
        for e in to_delete:
            db.session.delete(e)
    else:
        Timetable.query.filter_by(locked=False).delete()
    db.session.commit()

    allocations  = Allocation.query.all()
    faculty_map  = {f.id: f for f in Faculty.query.all()}
    subject_map  = {s.id: s for s in Subject.query.all()}
    division_map = {d.id: d for d in Division.query.all()}
    room_list    = Room.query.all()

    # Filter allocations to selected session only
    if session_type:
        sems = ODD_SEMESTERS if session_type == 'Odd' else EVEN_SEMESTERS
        allocations = [a for a in allocations if division_map[a.division_id].semester in sems]

    entries = generate_timetable(allocations, faculty_map, subject_map, division_map, room_list)
    for e in entries:
        db.session.add(Timetable(**e))
    db.session.commit()
    flash(f'Timetable generated ({session_type or "All"} session) — {len(entries)} entries.', 'success')
    return redirect(url_for('tt_bp.timetable_view', session_type=session_type))

@tt_bp.route('/timetable/lock/<int:eid>', methods=['POST'])
@login_required
def timetable_lock(eid):
    entry = Timetable.query.get_or_404(eid)
    entry.locked = not entry.locked
    db.session.commit()
    return jsonify({'locked': entry.locked})

@tt_bp.route('/timetable/export/excel')
@login_required
def export_excel():
    import io
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Timetable'
    ws.append(['Day', 'Slot', 'Faculty', 'Subject', 'Division', 'Room', 'Batch'])
    for e in Timetable.query.all():
        ws.append([
            e.day, e.slot,
            e.faculty.name  if e.faculty  else '',
            e.subject.subject_name if e.subject else '',
            f"{e.division.course} Sem{e.division.semester} {e.division.division}" if e.division else '',
            e.room.room_no  if e.room    else '',
            e.batch or '',
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='timetable.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@tt_bp.route('/timetable/export/pdf')
@login_required
def export_pdf():
    import io
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
    from reportlab.lib import colors
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
    data = [['Day', 'Slot', 'Faculty', 'Subject', 'Division', 'Room', 'Batch']]
    for e in Timetable.query.all():
        data.append([
            e.day, str(e.slot),
            e.faculty.name  if e.faculty  else '',
            e.subject.subject_name if e.subject else '',
            f"{e.division.course} Sem{e.division.semester} {e.division.division}" if e.division else '',
            e.room.room_no  if e.room    else '',
            e.batch or '',
        ])
    t = Table(data)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#222')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('GRID',       (0,0), (-1,-1), 0.5, colors.grey),
        ('FONTSIZE',   (0,0), (-1,-1), 8),
    ]))
    doc.build([t])
    buf.seek(0)
    return send_file(buf, download_name='timetable.pdf',
                     as_attachment=True, mimetype='application/pdf')
